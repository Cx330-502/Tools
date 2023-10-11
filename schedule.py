import os
import time
from datetime import datetime
from openpyxl import *
from itertools import combinations
import xlwt


class Printer:
    def __init__(self):
        self.current_ch = '-'
        self.progress0 = 0
        self.start_time = datetime.now()
        self.minutes = 0
        self.break_out = False

    def init_print(self):
        print("Welcome to use Cx330_502's schedule program, the program running time may be long, please be "
              "patient!")
        print("    ___  _  _  ___  ___   ___       ___   ___  ___  ")
        print("   / __)( \\/ )(__ )(__ ) / _ \\     | __) / _ \\(__ \\ ")
        print("  ( (__  )  (  (_ \\ (_ \\( (_) )___ |__ \\( (_) )/ _/ ")
        print("   \\___)(_/\\_)(___/(___/ \\___/(___)(___/ \\___/(____)")
        print()

    def update_time(self):
        time.sleep(0.5)
        self.current_ch = '-'
        self.start_time = datetime.now()
        self.progress0 = 0
        self.minutes = 0
        self.break_out = False

    def print_progress(self, progress_name, c):
        scale = 50
        progress = '█' * int(c // (100 / scale))
        point = '.' * int(scale - c // (100 / scale))
        during = (datetime.now() - self.start_time).total_seconds()
        if c == 100:
            self.current_ch = '>'
        else:
            if self.current_ch == '-':
                self.current_ch = '\\'
            elif self.current_ch == '\\':
                self.current_ch = '|'
            elif self.current_ch == '|':
                self.current_ch = '/'
            elif self.current_ch == '/':
                self.current_ch = '-'
        print("\r{} {}: {:^3.0f}%【{}->{}】{:.2f}s".format(progress_name, self.current_ch, c, progress, point, during),
              end="")

    def regular_print(self, progress_name, value):
        if (datetime.now() - self.start_time).seconds // value < 50:
            self.progress0 = (datetime.now() - self.start_time).seconds // value * 100 / 50
            self.print_progress(progress_name, printer.progress0)
        else:
            self.print_progress(progress_name, printer.progress0)
            if (datetime.now() - self.start_time).seconds // 60 > self.minutes:
                self.minutes = (datetime.now() - self.start_time).seconds // 60
                self.middle_print(progress_name)
                if self.minutes % 5 == 0:
                    self.break_print(progress_name)

    def end_print(self, progress_name):
        self.progress0 = 100
        self.print_progress(progress_name, printer.progress0)
        print()

    def middle_print(self, progress_name):
        print()
        print()
        print("The scheduling process has taken long due to data issues or some other reasons.")
        time0 = round((datetime.now() - self.start_time).total_seconds(), 2)
        time.sleep(0.5)
        print("It has been waiting for " + str(time0) +
              " seconds in process " + progress_name +
              " and is still calculating the optimal scheduling process.")
        time.sleep(0.5)
        if self.minutes == 1:
            target = ""
            if progress_name == "Schedule1":
                target = "Schedule 1 is to prioritize the scheduling of time slots with fewer people to choose from."
            elif progress_name == "Schedule2":
                target = "Schedule 2 is to get the remaining time slots filled as much as possible."
            elif progress_name == "Schedule3":
                target = "Schedule 3 is to get more scheduling for those with less scheduling."
            elif progress_name == "Schedule4":
                target = "Schedule 4 is designed to allow those with more shifts to get fewer shifts."
            elif progress_name == "Schedule5":
                target = "Schedule 5 is designed to allow those with more shifts to get fewer shifts."
            print(target)
            time.sleep(0.5)
        print("I will then output a version of the current schedule, and you can choose to use it or wait.")

        workbook0 = xlwt.Workbook()
        # 创建一个样式
        style0 = xlwt.XFStyle()
        # 设置单元格文本自动换行
        alignment0 = xlwt.Alignment()
        alignment0.wrap = 1
        alignment0.horz = xlwt.Alignment.HORZ_CENTER
        alignment0.vert = xlwt.Alignment.VERT_CENTER
        style0.alignment = alignment0

        worksheet = workbook0.add_sheet('值日表')
        for i in range(len(dates)):
            worksheet.col(i).width = 4000
            worksheet.write(0, i + 1, dates[i], style)
        for j in range(len(times)):
            worksheet.row(j).height = 200
            worksheet.write(j + 1, 0, times[j], style)
        for i in range(len(dates)):
            for j in range(len(times)):
                names = ""
                for key, value in table[i]['times'][j]['people'].items():
                    names = names + key + '\n'
                worksheet.write(j + 1, i + 1, names, style)

        worksheet = workbook0.add_sheet('每人安排')
        for i in range(10):
            worksheet.col(i).width = 4000
        worksheet.row(0).height = 200
        worksheet.write(0, 0, "姓名", style)
        worksheet.write(0, 1, "身份", style)
        worksheet.write(0, 2, "值班次数", style)
        worksheet.write_merge(0, 0, 3, 9, "值班时间", style)
        for j in range(len(people_0)):
            worksheet.row(j + 1).height = 200
            worksheet.write(j + 1, 0, people_0[j].name, style)
            worksheet.write(j + 1, 1, people_0[j].status, style)
            worksheet.write(j + 1, 2, len(people_0[j].sche_time), style)
            for i in range(len(people_0[j].sche_time)):
                worksheet.write(j + 1, i + 3, people_0[j].sche_time[i], style)
        name = "TempOutput" + str(datetime.now().strftime("%H%M%S")) + ".xlsx"
        workbook0.save(os.path.join(output_root, name))

        print("The intermediate version is now exported to " + name + " , so check it out and keep waiting!")
        print()

    def break_print(self, progress_name):
        print()
        print()
        print("The " + progress_name + " process has taken more than " + str(self.minutes) + " minutes.")
        time.sleep(0.2)
        print("Are you sure to wait for the process to complete?")
        time.sleep(0.2)
        print("If you choose to wait, the program will continue to run " + progress_name + ". And if you choose not to "
                                                                                           "wait the program will run next process, which may make the schedule not that perfect.")
        time.sleep(0.2)
        while True:
            temp = input("Please enter Y to continue, or enter N to skip : ")
            if temp == "N" or temp == "n":
                print()
                self.break_out = True
                return
            elif temp == "Y" or temp == "y":
                print()
                return


class People:
    name = ""
    status = ""
    free_time = []
    sche_time = []
    compensation = 0

    def __init__(self, name, status):
        self.name = name
        self.status = status
        self.free_time = []
        self.sche_time = []
        self.compensation = 0


class Sche_Time:
    def __init__(self, printer0):
        self.printer = printer0

    def get_times(self):
        column_to_read = sheet1['A']  # 读取第一列
        for cell in column_to_read:
            if cell.value is None or cell.value == '值班时段':
                continue
            else:
                times.append(cell.value)
        column_to_read = sheet2['A']  # 读取第一列
        for cell in column_to_read:
            if cell.value is None or cell.value in times or cell.value == '值班时段':
                continue
            else:
                times.append(cell.value)

    def get_dates(self):
        row_to_read = sheet1[1]  # 读取第一行
        for cell in row_to_read:
            if cell.value is None or cell.value == '值班时段' or cell.value == '姓名':
                continue
            else:
                dates.append(cell.value)
        row_to_read = sheet2[1]  # 读取第一行
        for cell in row_to_read:
            if cell.value is None or cell.value in dates or cell.value == '值班时段' or cell.value == '姓名':
                continue
            else:
                dates.append(cell.value)
        return dates

    def get_people(self):
        column_to_read = sheet1['B']  # 读取第二列
        for cell in column_to_read:
            if cell.value is None or cell.value == '姓名':
                continue
            else:
                people = People(cell.value, '助理')
                try:
                    people_1[cell.value]
                except KeyError:
                    people_1[cell.value] = people
        column_to_read = sheet2['B']  # 读取第二列
        for cell in column_to_read:
            if cell.value is None or cell.value == '姓名':
                continue
            else:
                people = People(cell.value, '导员')
                try:
                    people_2[cell.value]
                except KeyError:
                    people_2[cell.value] = people
        for key, value in people_1.items():
            people_0.append(value)
        for key, value in people_2.items():
            people_0.append(value)

    def get_table(self):
        for i in dates:
            table_item = {'date': i, 'times': []}
            for j in times:
                table_item['times'].append({'time': j, 'people': {}})
            table.append(table_item)
        return table

    def get_free_time(self):
        time = ""
        max_row = sheet1.max_row
        max_col = sheet1.max_column
        for row in sheet1.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
            name = ""
            i = 0
            for cell in row:
                if cell.value in times:
                    time = cell.value
                    continue
                elif cell.value is not None and name == "":
                    i = -1
                    name = cell.value
                    continue
                i = i + 1
                if cell.value is not None and name != "" and i < len(dates):
                    people_1[name].free_time.append({'date': dates[i], 'time': time})
        time = ""
        for row in sheet2.iter_rows(min_row=2):
            name = ""
            i = 0
            for cell in row:
                if cell.value in times:
                    time = cell.value
                elif cell.value is not None and cell.value != name:
                    i = -1
                    name = cell.value
                    continue
                i = i + 1
                if cell.value is not None and cell.value == name:
                    people_2[name].free_time.append({'date': dates[i], 'time': time})
        return people_1, people_2

    def get_free_people_list(self, i, j):
        free_people_list = []
        for key, values in people_1.items():
            if len(values.sche_time) > 5:
                continue
            if dates[i] + ' ' + times[j] in values.sche_time:
                continue
            for free_time in values.free_time:
                if free_time['date'] == dates[i] and free_time['time'] == times[j]:
                    free_people_list.append(values)
        for key, values in people_2.items():
            if len(values.sche_time) > 3:
                continue
            if dates[i] + ' ' + times[j] in values.sche_time:
                continue
            for free_time in values.free_time:
                if free_time['date'] == dates[i] and free_time['time'] == times[j]:
                    free_people_list.append(values)
        return free_people_list

    def schedule1(self):
        if printer.break_out:
            return True
        printer.regular_print("Schedule1", 0.005)
        for i in range(len(dates)):
            for j in range(len(times)):
                free_people_list = self.get_free_people_list(i, j)
                if len(free_people_list) + len(table[i]['times'][j]['people']) < 2:
                    return False
                if len(free_people_list) + len(table[i]['times'][j]['people']) <= 3:
                    for item in free_people_list:
                        try:
                            table[i]['times'][j]['people'][item.name]
                        except KeyError:
                            item.sche_time.append(dates[i] + ' ' + times[j])
                            table[i]['times'][j]['people'][item.name] = item

    def schedule2(self, i, j):
        if printer.break_out:
            return True
        printer.regular_print("Schedule2", 0.8)
        free_people_list = self.get_free_people_list(i, j)
        return_value = False
        min0 = 2 - len(table[i]['times'][j]['people'])
        if len(free_people_list) < min0:
            return False
        if min0 <= 0:
            if j != len(times) - 1:
                return_value = self.schedule2(i, j + 1)
            else:
                if i != len(dates) - 1:
                    return_value = self.schedule2(i + 1, 0)
                else:
                    return_value = True
        else:
            new_list = list(combinations(free_people_list, min0))
            length = len(new_list)
            temp = 0
            new_list2 = []
            for item in new_list:
                temp = temp + 1
                if item[0].compensation < 0 or item[1].compensation < 0:
                    new_list2.append(item)
                    continue
                for item0 in item:
                    item0.sche_time.append(dates[i] + ' ' + times[j])
                    table[i]['times'][j]['people'][item0.name] = item0
                if j != len(times) - 1:
                    return_value = self.schedule2(i, j + 1)
                else:
                    if i != len(dates) - 1:
                        return_value = self.schedule2(i + 1, 0)
                    else:
                        return_value = True
                if not return_value:
                    for item0 in item:
                        item0.sche_time.pop()
                        table[i]['times'][j]['people'].pop(item0.name)
                if return_value:
                    break
        if not return_value:
            return False
        else:
            return True

    def schedule3(self, i):
        if printer.break_out:
            return True
        printer.regular_print("Schedule3", 1)
        people = people_0[i]
        return_value = False
        if (people.status == '导员' and len(people.sche_time) > 3) or (
                people.status == '助理' and len(people.sche_time) > 4):
            if i == len(people_0) - 1:
                return True
            else:
                return self.schedule3(i + 1)
        for free_time in people.free_time:
            if free_time['date'] + ' ' + free_time['time'] in people.sche_time:
                continue
            for item in table:
                if item['date'] == free_time['date']:
                    for item2 in item['times']:
                        if item2['time'] == free_time['time']:
                            if len(item2['people']) > 4:
                                break
                            item2['people'][people.name] = people
                            people.sche_time.append(free_time['date'] + ' ' + free_time['time'])
                            if (people.status == '导员' and len(people.sche_time) >= 3) or (
                                    people.status == '助理' and len(people.sche_time) >= 4):
                                if i == len(people_0) - 1:
                                    return True
                                return_value = self.schedule3(i + 1)
                                if return_value:
                                    return True
                                else:
                                    people.sche_time.pop()
                                    item2['people'].pop(people.name)
                                    break
        return False

    def schedule4(self):
        for people in people_0:
            if printer.break_out:
                return True
            printer.regular_print("Schedule4", 0.5)
            if (people.status == '导员' and len(people.sche_time) >= 3) or (
                    people.status == '助理' and len(people.sche_time) >= 4):
                continue
            free_time_list = []
            for free_time in people.free_time:
                if free_time['date'] + ' ' + free_time['time'] not in people.sche_time:
                    free_time_list.append(free_time)
            for free_time in free_time_list:
                for people0 in people_0:
                    if (people0.status == '导员' and len(people0.sche_time) > 3) or (
                            people0.status == '助理' and len(people0.sche_time) > 4):
                        if free_time['date'] + ' ' + free_time['time'] in people0.sche_time:
                            people0.sche_time.remove(free_time['date'] + ' ' + free_time['time'])
                            for item in table:
                                if item['date'] == free_time['date']:
                                    for item2 in item['times']:
                                        if item2['time'] == free_time['time']:
                                            del item2['people'][people0.name]
                                            item2['people'][people.name] = people
                            break
                if (people.status == '导员' and len(people.sche_time) >= 3) or (
                        people.status == '助理' and len(people.sche_time) >= 4):
                    break

    def schedule5(self):
        for people in people_0:
            if printer.break_out:
                return True
            printer.regular_print("Schedule5", 0.5)
            if (people.status == '导员' and len(people.sche_time) > 3) or (
                    people.status == '助理' and len(people.sche_time) > 4):
                continue
            free_time_list = []
            for free_time in people.free_time:
                if free_time['date'] + ' ' + free_time['time'] not in people.sche_time:
                    free_time_list.append(free_time)
            for free_time in free_time_list:
                for people0 in people_0:
                    if (people0.status == '导员' and len(people0.sche_time) > 3) or (
                            people0.status == '助理' and len(people0.sche_time) > 4):
                        if free_time['date'] + ' ' + free_time['time'] in people0.sche_time:
                            people0.sche_time.remove(free_time['date'] + ' ' + free_time['time'])
                            for item in table:
                                if item['date'] == free_time['date']:
                                    for item2 in item['times']:
                                        if item2['time'] == free_time['time']:
                                            del item2['people'][people0.name]
                                            item2['people'][people.name] = people
                            break
                if (people.status == '导员' and len(people.sche_time) >= 3) or (
                        people.status == '助理' and len(people.sche_time) >= 4):
                    break

    def export_sche(self):
        worksheet = workbook.add_sheet('值日表')
        for i in range(len(dates)):
            worksheet.col(i).width = 4000
            worksheet.write(0, i + 1, dates[i], style)
        for j in range(len(times)):
            worksheet.row(j).height = 200
            worksheet.write(j + 1, 0, times[j], style)
        for i in range(len(dates)):
            for j in range(len(times)):
                names = ""
                for key, value in table[i]['times'][j]['people'].items():
                    names = names + key + '\n'
                worksheet.write(j + 1, i + 1, names, style)

    def export_people(self):
        worksheet = workbook.add_sheet('每人安排')
        for i in range(10):
            worksheet.col(i).width = 4000
        worksheet.row(0).height = 200
        worksheet.write(0, 0, "姓名", style)
        worksheet.write(0, 1, "身份", style)
        worksheet.write(0, 2, "值班次数", style)
        worksheet.write_merge(0, 0, 3, 9, "值班时间", style)
        for j in range(len(people_0)):
            worksheet.row(j + 1).height = 200
            worksheet.write(j + 1, 0, people_0[j].name, style)
            worksheet.write(j + 1, 1, people_0[j].status, style)
            worksheet.write(j + 1, 2, len(people_0[j].sche_time), style)
            for i in range(len(people_0[j].sche_time)):
                worksheet.write(j + 1, i + 3, people_0[j].sche_time[i], style)
        workbook.save(os.path.join(output_root, "output.xlsx"))


def input_model():
    print("Notice that the input file must be named 'sche.xlsx' ")
    print("Please input the model you want to use:")
    print("1. input and output both in current directory")
    print("2. './data/schedule' for input and './output/schedule' for output")
    print("3. Customizing the working directory")
    while True:
        model = input("Please input 1 or 2 or 3: ")
        model = int(model)
        if model == 1:
            input_root0 = "./"
            output_root0 = "./"
            break
        elif model == 2:
            input_root0 = "./data/schedule"
            output_root0 = "./output/schedule"
            break
        elif model == 3:
            input_root0 = input("Please input the input directory: ")
            output_root0 = input("Please input the output directory: ")
            break
    os.makedirs(input_root0, exist_ok=True)
    os.makedirs(output_root0, exist_ok=True)
    return input_root0, output_root0


if __name__ == '__main__':
    printer = Printer()
    printer.init_print()
    input_root, output_root = input_model()
    
    times = []
    dates = []
    people_1 = {}
    people_2 = {}
    people_0 = []
    table = []
    workbook = xlwt.Workbook()
    # 创建一个样式
    style = xlwt.XFStyle()
    # 设置单元格文本自动换行
    alignment = xlwt.Alignment()
    alignment.wrap = 1
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    alignment.vert = xlwt.Alignment.VERT_CENTER
    style.alignment = alignment

    wb = load_workbook(os.path.join(input_root, "input.xlsx"))  # 打开excel文件
    sheet_names = wb.sheetnames  # 获取工作表名
    if len(sheet_names) != 2:
        print("工作表数量不对")
        exit(0)
    sheet1 = wb[sheet_names[0]]  # 助理工作表
    sheet2 = wb[sheet_names[1]]  # 导员工作表

    
    sche_time = Sche_Time(printer)

    printer.update_time()
    printer.print_progress("Preparations", 0)
    sche_time.get_times()
    printer.print_progress("Preparations", 20)
    sche_time.get_dates()
    printer.print_progress("Preparations", 40)
    sche_time.get_people()
    printer.print_progress("Preparations", 60)
    sche_time.get_table()
    printer.print_progress("Preparations", 80)
    sche_time.get_free_time()
    printer.print_progress("Preparations", 100)
    print()

    printer.update_time()
    sche_time.schedule1()
    printer.end_print("Schedule1")

    printer.update_time()
    sche_time.schedule2(0, 0)
    printer.end_print("Schedule2")

    printer.update_time()
    sche_time.schedule3(0)
    printer.end_print("Schedule3")

    printer.update_time()
    sche_time.schedule4()
    printer.end_print("Schedule4")

    printer.update_time()
    sche_time.schedule5()
    printer.end_print("Schedule5")

    printer.update_time()
    sche_time.export_sche()
    printer.end_print("Export1")

    printer.update_time()
    sche_time.export_people()
    printer.end_print("Export2")
