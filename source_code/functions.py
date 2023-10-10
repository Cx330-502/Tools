from openpyxl import *
from models import *
from itertools import combinations
import xlwt

times = []
dates = []
people_1 = {}
people_2 = {}
people_0 = []
table = []
workbook = xlwt.Workbook()
# 创建一个样式
style = xlwt.XFStyle()
# 设置单元格文本自动换行
alignment = xlwt.Alignment()
alignment.wrap = 1
alignment.horz = xlwt.Alignment.HORZ_CENTER
alignment.vert = xlwt.Alignment.VERT_CENTER
style.alignment = alignment

wb = load_workbook('sche.xlsx')
sheet_names = wb.sheetnames  # 获取工作表名
if len(sheet_names) != 2:
    print("工作表数量不对")
    exit(0)
sheet1 = wb[sheet_names[0]]  # 助理工作表
sheet2 = wb[sheet_names[1]]  # 导员工作表


def get_times():
    column_to_read = sheet1['A']  # 读取第一列
    for cell in column_to_read:
        if cell.value is None or cell.value == '值班时段':
            continue
        else:
            times.append(cell.value)
    column_to_read = sheet2['A']  # 读取第一列
    for cell in column_to_read:
        if cell.value is None or cell.value in times or cell.value == '值班时段':
            continue
        else:
            times.append(cell.value)


def get_dates():
    row_to_read = sheet1[1]  # 读取第一行
    for cell in row_to_read:
        if cell.value is None or cell.value == '值班时段' or cell.value == '姓名':
            continue
        else:
            dates.append(cell.value)
    row_to_read = sheet2[1]  # 读取第一行
    for cell in row_to_read:
        if cell.value is None or cell.value in dates or cell.value == '值班时段' or cell.value == '姓名':
            continue
        else:
            dates.append(cell.value)
    return dates


def get_people():
    column_to_read = sheet1['B']  # 读取第二列
    for cell in column_to_read:
        if cell.value is None or cell.value == '姓名':
            continue
        else:
            people = People(cell.value, '助理')
            try:
                people_1[cell.value]
            except KeyError:
                people_1[cell.value] = people
    column_to_read = sheet2['B']  # 读取第二列
    for cell in column_to_read:
        if cell.value is None or cell.value == '姓名':
            continue
        else:
            people = People(cell.value, '导员')
            try:
                people_2[cell.value]
            except KeyError:
                people_2[cell.value] = people
    for key, value in people_1.items():
        people_0.append(value)
    for key, value in people_2.items():
        people_0.append(value)


def get_table():
    for i in dates:
        table_item = {'date': i, 'times': []}
        for j in times:
            table_item['times'].append({'time': j, 'people': {}})
        table.append(table_item)
    return table


def get_free_time():
    time = ""
    max_row = sheet1.max_row
    max_col = sheet1.max_column
    for row in sheet1.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        name = ""
        i = 0
        for cell in row:
            if cell.value in times:
                time = cell.value
                continue
            elif cell.value is not None and name == "":
                i = -1
                name = cell.value
                continue
            i = i + 1
            if cell.value is not None and name != "" and i < len(dates):
                people_1[name].free_time.append({'date': dates[i], 'time': time})
    time = ""
    for row in sheet2.iter_rows(min_row=2):
        name = ""
        i = 0
        for cell in row:
            if cell.value in times:
                time = cell.value
            elif cell.value is not None and cell.value != name:
                i = -1
                name = cell.value
                continue
            i = i + 1
            if cell.value is not None and cell.value == name:
                people_2[name].free_time.append({'date': dates[i], 'time': time})
    return people_1, people_2


def get_free_people_list(i, j):
    free_people_list = []
    for key, values in people_1.items():
        if len(values.sche_time) > 5:
            continue
        if dates[i] + ' ' + times[j] in values.sche_time:
            continue
        for free_time in values.free_time:
            if free_time['date'] == dates[i] and free_time['time'] == times[j]:
                free_people_list.append(values)
    for key, values in people_2.items():
        if len(values.sche_time) > 3:
            continue
        if dates[i] + ' ' + times[j] in values.sche_time:
            continue
        for free_time in values.free_time:
            if free_time['date'] == dates[i] and free_time['time'] == times[j]:
                free_people_list.append(values)
    return free_people_list


def schedule1():
    for i in range(len(dates)):
        for j in range(len(times)):
            free_people_list = get_free_people_list(i, j)
            if len(free_people_list) + len(table[i]['times'][j]['people']) < 2:
                return False
            if len(free_people_list) + len(table[i]['times'][j]['people']) <= 3:
                for item in free_people_list:
                    try:
                        table[i]['times'][j]['people'][item.name]
                    except KeyError:
                        item.sche_time.append(dates[i] + ' ' + times[j])
                        table[i]['times'][j]['people'][item.name] = item


def schedule2(i, j):
    free_people_list = get_free_people_list(i, j)
    return_value = False
    min0 = 2 - len(table[i]['times'][j]['people'])
    if len(free_people_list) < min0:
        return False
    if min0 <= 0:
        if j != len(times) - 1:
            return_value = schedule2(i, j + 1)
        else:
            if i != len(dates) - 1:
                return_value = schedule2(i + 1, 0)
            else:
                return_value = True
    else:
        new_list = list(combinations(free_people_list, min0))
        length = len(new_list)
        temp = 0
        new_list2 = []
        for item in new_list:
            temp = temp + 1
            if item[0].compensation < 0 or item[1].compensation < 0:
                new_list2.append(item)
                continue
            for item0 in item:
                item0.sche_time.append(dates[i] + ' ' + times[j])
                table[i]['times'][j]['people'][item0.name] = item0
            if j != len(times) - 1:
                return_value = schedule2(i, j + 1)
            else:
                if i != len(dates) - 1:
                    return_value = schedule2(i + 1, 0)
                else:
                    return_value = True
            if not return_value:
                for item0 in item:
                    item0.sche_time.pop()
                    table[i]['times'][j]['people'].pop(item0.name)
            if return_value:
                break
    if not return_value:
        return False
    else:
        return True


def schedule3(i):
    people = people_0[i]
    return_value = False
    if (people.status == '导员' and len(people.sche_time) > 3) or (
            people.status == '助理' and len(people.sche_time) > 4):
        if i == len(people_0) - 1:
            return True
        else:
            return schedule3(i + 1)
    for free_time in people.free_time:
        if free_time['date'] + ' ' + free_time['time'] in people.sche_time:
            continue
        for item in table:
            if item['date'] == free_time['date']:
                for item2 in item['times']:
                    if item2['time'] == free_time['time']:
                        if len(item2['people']) > 4:
                            break
                        item2['people'][people.name] = people
                        people.sche_time.append(free_time['date'] + ' ' + free_time['time'])
                        if (people.status == '导员' and len(people.sche_time) >= 3) or (
                                people.status == '助理' and len(people.sche_time) >= 4):
                            if i == len(people_0) - 1:
                                return True
                            return_value = schedule3(i + 1)
                            if return_value:
                                return True
                            else:
                                people.sche_time.pop()
                                item2['people'].pop(people.name)
                                break
    return False


def schedule4():
    for people in people_0:
        if (people.status == '导员' and len(people.sche_time) >= 3) or (
                people.status == '助理' and len(people.sche_time) >= 4):
            continue
        free_time_list = []
        for free_time in people.free_time:
            if free_time['date'] + ' ' + free_time['time'] not in people.sche_time:
                free_time_list.append(free_time)
        for free_time in free_time_list:
            for people0 in people_0:
                if (people0.status == '导员' and len(people0.sche_time) > 3) or (
                        people0.status == '助理' and len(people0.sche_time) > 4):
                    if free_time['date'] + ' ' + free_time['time'] in people0.sche_time:
                        people0.sche_time.remove(free_time['date'] + ' ' + free_time['time'])
                        for item in table:
                            if item['date'] == free_time['date']:
                                for item2 in item['times']:
                                    if item2['time'] == free_time['time']:
                                        del item2['people'][people0.name]
                                        item2['people'][people.name] = people
                        break
            if (people.status == '导员' and len(people.sche_time) >= 3) or (
                    people.status == '助理' and len(people.sche_time) >= 4):
                break


def schedule5():
    for people in people_0:
        if (people.status == '导员' and len(people.sche_time) > 3) or (
                people.status == '助理' and len(people.sche_time) > 4):
            continue
        free_time_list = []
        for free_time in people.free_time:
            if free_time['date'] + ' ' + free_time['time'] not in people.sche_time:
                free_time_list.append(free_time)
        for free_time in free_time_list:
            for people0 in people_0:
                if (people0.status == '导员' and len(people0.sche_time) > 3) or (
                        people0.status == '助理' and len(people0.sche_time) > 4):
                    if free_time['date'] + ' ' + free_time['time'] in people0.sche_time:
                        people0.sche_time.remove(free_time['date'] + ' ' + free_time['time'])
                        for item in table:
                            if item['date'] == free_time['date']:
                                for item2 in item['times']:
                                    if item2['time'] == free_time['time']:
                                        del item2['people'][people0.name]
                                        item2['people'][people.name] = people
                        break
            if (people.status == '导员' and len(people.sche_time) >= 3) or (
                    people.status == '助理' and len(people.sche_time) >= 4):
                break


def export_sche():
    worksheet = workbook.add_sheet('值日表')
    for i in range(len(dates)):
        worksheet.col(i).width = 4000
        worksheet.write(0, i + 1, dates[i], style)
    for j in range(len(times)):
        worksheet.row(j).height = 200
        worksheet.write(j + 1, 0, times[j], style)
    for i in range(len(dates)):
        for j in range(len(times)):
            names = ""
            for key, value in table[i]['times'][j]['people'].items():
                names = names + key + '\n'
            worksheet.write(j + 1, i + 1, names, style)


def export_people():
    worksheet = workbook.add_sheet('每人安排')
    for i in range(10):
        worksheet.col(i).width = 4000
    worksheet.row(0).height = 200
    worksheet.write(0, 0, "姓名", style)
    worksheet.write(0, 1, "身份", style)
    worksheet.write(0, 2, "值班次数", style)
    worksheet.write_merge(0, 0, 3, 9, "值班时间", style)
    for j in range(len(people_0)):
        worksheet.row(j + 1).height = 200
        worksheet.write(j+1, 0, people_0[j].name, style)
        worksheet.write(j+1, 1, people_0[j].status, style)
        worksheet.write(j+1, 2, len(people_0[j].sche_time), style)
        for i in range(len(people_0[j].sche_time)):
            worksheet.write(j+1, i+3, people_0[j].sche_time[i], style)
    workbook.save("222.xlsx")
